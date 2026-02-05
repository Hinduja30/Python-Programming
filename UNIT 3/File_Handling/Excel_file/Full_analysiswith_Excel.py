import os
import openpyxl

base = os.path.dirname(__file__)
path = os.path.join(base , 'VideoSales.xlsx')
wb = openpyxl.load_workbook(path)
print(type(wb))

ws = wb.active
print(type(ws))
print(ws)

ws1 = wb['A New Sheet1']
print(ws1)

print(f'Total number of rows{ws.max_row} and total number of columns {ws.max_column}')
print(f'the value in cell A1 is : {ws['A1'].value}')

data=[ws.cell(row=i,column=2).value for i in range(1 , ws.max_column+1)]
print(data)

data1 = [ws.cell(row=i,column=2).value for i in range(2,12)]
print(data1)

mylist = list()
for value in ws.iter_rows(min_row=1,max_row=11,min_col=1,max_col=4,values_only=True):
    mylist.append(value)

for ele1,ele2,ele3,ele4 in mylist:
    print("{:<35}{:<8}{:<10}{:<10}".format(ele1,ele2,ele3,ele4))

#writing to cell
ws['K1']='Sum of sales'

#OR
#ws.cell(row=1,column=11,value='Sum of sales')
wb.save('VideoSales_1.xlsx')

#adding new column
#total_sales = ((ws.cell(row=2,column=7))+(ws.cell(row=2,column=8))+(ws.cell(row=2,column=9))+(ws.cell(row=2,column=10)))

#ws.cell(row=2,column=11).value=total_sales   this dont work as python canth add objects
#wb.save('VideoSales_1.xlsx')

for i in range(1,ws.max_row):
    NA_sales=ws.cell(row=1,column=7).value
    EU_sales=ws.cell(row=1,column=8).value
    JP_sales=ws.cell(row=1,column=9).value
    Other_sales=ws.cell(row=1,column=10).value

    total_sales = (NA_sales+EU_sales+JP_sales+Other_sales)
    ws.cell(row=1,column=11).value=total_sales
wb.save('VideoSales_1.xlsx')

#adding new row
new_row=(31,'Crichet Premier league','PC',2024,'Sports','GameStudio India',2.50,1.80,0.90,1.20,6.40)
ws.append(new_row)
    
wb.save("VideoSales_1.xlsx")
values=[ws.cell(row=ws.max_row,column=i) for i in range(1, ws.max_column+1)]
print(values)

#delete row
ws.delete_rows(ws.max_row,1)
wb.save('VideoSales_2.xlsx')

#Excel formulas
ws['P1']="Average Sales"
ws['P2']= "=AVERAGE(K2:K31)"
ws['Q1']="Number of populated cells"
ws['Q2']="=COUNTA(K2:K31)"
ws['R1']="Number of rows with sports genre"
ws['R2']='=COUNTIF(E2:E32,"Sports" )'
ws['S1']="Total sports Sales"
ws['S2']='=SUMIF(E2:E31,"Sports",K2:K31)'
ws['T1']="Rounded sum of Sports Sales"
ws['T2']="=CEILING(S2,S5)"

print(ws.title)

ws.title='Games sales data'

wb.save('VideoSales_formulas.xlsx')

print(ws.title)

print(wb.active.title)

#create delete and duplicate new worksheet
wb.create_sheet('Empty sheet')
print(wb.sheetnames)
wb.remove(wb['Empty sheet'])
print(wb.sheetnames)
wb.copy_worksheet(wb['Games sales data'])
wb.save('VideoSales_Formula_2.xlsx')
print(wb.sheetnames)







